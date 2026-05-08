import os
import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy
from django.conf import settings
from ..models import Compra
from compras.utils.string_utils import StringUtils

class ExcelConferenciaService:
    @staticmethod
    def _copy_column_style(ws, source_col_idx, target_col_idx):
        """Helper to copy style and width from one column to another."""
        source_letter = get_column_letter(source_col_idx)
        target_letter = get_column_letter(target_col_idx)
        
        # Copy Width
        if source_letter in ws.column_dimensions:
            ws.column_dimensions[target_letter].width = ws.column_dimensions[source_letter].width

        # Copy cell styles for a reasonable number of rows (e.g., 200)
        for row in range(1, 201):
            source_cell = ws.cell(row=row, column=source_col_idx)
            target_cell = ws.cell(row=row, column=target_col_idx)
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

    @staticmethod
    def generate_excel(compra_id):
        # Prefetch demandas, itens, and pesquisas to avoid N+1 queries
        compra = Compra.objects.prefetch_related('demandas__itens__pesquisas').get(pk=compra_id)
        
        template_dir = os.path.join(settings.BASE_DIR, 'compras', 'templates_docs')
        template_path = os.path.join(template_dir, "CONFERENCIA_LICITACAO_MODELO_C.xlsx")
        
        if not os.path.exists(template_path):
            # Fallback to old template name if the new one doesn't exist
            template_path = os.path.join(template_dir, "CONFERENCIA_LICITACAO_MODELO.xlsx")
            if not os.path.exists(template_path):
                return None
            
        # Load template
        wb = load_workbook(template_path)
        # Enable full calculation on load to refresh formulas like UNIQUE()
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = 'auto'
        ws = wb.active
        
        # Mapping companies to columns
        # Column F (6) is COMPRASGOVBR
        # Columns G (7) onwards are reserved for up to 50 other companies
        fixed_companies_start_col = 7
        total_reserved_cols = 50
        
        # 1. Identify all unique other companies and check if ComprasGov is used
        other_companies = []
        compras_gov_used = False
        for demanda in compra.demandas.all():
            for item in demanda.itens.all():
                for pesquisa in item.pesquisas.all():
                    name = (pesquisa.nome_fornecedor or "").strip().upper()
                    if not name:
                        continue
                    if "COMPRASGOVBR" in name:
                        compras_gov_used = True
                        continue
                    if name not in other_companies:
                        other_companies.append(name)
        
        # 2. Write company headers (starting from G)
        for i, company_name in enumerate(other_companies[:total_reserved_cols]):
            ws.cell(row=4, column=fixed_companies_start_col + i, value=company_name)

        # 4. Write data rows starting from row 5
        current_row = 5
        for demanda in compra.demandas.all():
            for item in demanda.itens.all():
                # A: Item
                ws.cell(row=current_row, column=1, value=item.numero_ordem)
                # B: Grupo Orçamentário
                ws.cell(row=current_row, column=2, value=demanda.grupo_orcamentario)
                # C: Elemento de despesa (Item de despesa)
                ws.cell(row=current_row, column=3, value=item.item_despesa)
                # D: DFD (Demanda)
                ws.cell(row=current_row, column=4, value= StringUtils.formatar_numero_demanda(demanda.numero_demanda))
                # E: CATMAT (Material code)
                ws.cell(row=current_row, column=5, value=item.codigo_material)
                
                # Fill Research Values
                for pesquisa in item.pesquisas.all():
                    name = (pesquisa.nome_fornecedor or "").strip().upper()
                    valor = pesquisa.valor_unitario
                    if not name:
                        continue
                        
                    if "COMPRASGOVBR" in name:
                        # F: COMPRASGOVBR
                        ws.cell(row=current_row, column=6, value=valor)
                    elif name in other_companies:
                        # G onwards: Other Companies
                        idx = other_companies.index(name)
                        if idx < total_reserved_cols:
                            ws.cell(row=current_row, column=fixed_companies_start_col + idx, value=valor)
                
                # Quantity: Column BT (72)
                qty_col = 72
                ws.cell(row=current_row, column=qty_col, value=item.quantidade)
                
                current_row += 1
                
        # 5. Hide unused rows from the last filled row up to row 216
        for row in range(current_row, 217):
            ws.row_dimensions[row].hidden = True

        # 6. Handle column visibility (F to BJ) at the end to ensure it's effective
        # Column F visibility
        ws.column_dimensions['F'].hidden = not compras_gov_used
        
        # Columns G (7) to BJ (62)
        total_other_companies = len(other_companies)
        for col_idx in range(fixed_companies_start_col, 63): # 7 to 62
            col_letter = get_column_letter(col_idx)
            # Company columns are 7 to 7 + total_reserved_cols - 1 (7 to 56)
            if fixed_companies_start_col <= col_idx < (fixed_companies_start_col + total_reserved_cols):
                idx = col_idx - fixed_companies_start_col
                if idx < total_other_companies:
                    ws.column_dimensions[col_letter].hidden = False
                else:
                    ws.column_dimensions[col_letter].hidden = True
            else:
                # For any other columns in the F-BJ range that are NOT company slots
                # (e.g., 57 to 62), we hide them as they are "not filled companies" per user request
                ws.column_dimensions[col_letter].hidden = True

        # 7. Clear BX5:BY216 and set UNIQUE formula in BX5
        # BX is 76, BY is 77
        for row in range(5, 217):
            ws.cell(row=row, column=76).value = None
            ws.cell(row=row, column=77).value = None
        ws.cell(row=5, column=76, value="=_xlfn.UNIQUE(B5:C216)")

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

